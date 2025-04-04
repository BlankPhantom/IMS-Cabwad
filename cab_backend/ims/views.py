import logging
import tempfile
from django.utils import timezone
from django.utils.timezone import now
from django.http import JsonResponse
from django.db.models import Sum, Q
from rest_framework.pagination import PageNumberPagination
logger = logging.getLogger(__name__)
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authentication import SessionAuthentication, TokenAuthentication 
from rest_framework.permissions import IsAuthenticated, BasePermission, IsAdminUser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authtoken.models import Token
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from ims.models import (Item, BeginningBalance, Classification, Measurement, Section, Purpose, TransactionProduct, 
                        TransactionDetails,RunningBalance, Area, MonthlyConsumption, MonthlyConsumptionTotal )                     
from ims.serializers import (ItemListSerializer, UserSerializer,ItemSerializer, BeginningBalanceSerializer, ClassificationSerializer, MeasurementSerializer, 
                             SectionSerializer, PurposeSerializer, TransactionProductSerializer, TransactionDetailsSerializer, 
                             RunningBalanceSerializer, AreaSerializer, MonthlyConsumptionTotalSerializer, MonthlyConsumptionSerializer) 
class IsSuperAdmin(BasePermission):
    """Allows access only to superadmins."""
    def has_permission(self, request, view):
        return request.user and request.user.is_superuser
class Pagination(PageNumberPagination):
    page_size = 20  # Load 20 items per request
    page_size_query_param = 'page_size'
    max_page_size = 100
    
@api_view(['POST'])
@authentication_classes([])
def login(request):
    username = request.data.get("username")
    password = request.data.get("password")

    if user := authenticate(username=username, password=password):
        token, created = Token.objects.get_or_create(user=user)

        return Response({
            "token": token.key,
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "is_superuser": user.is_superuser,  # ✅ Ensure this is included
                "is_staff": user.is_staff  # ✅ Ensure this is included
            }
        }, status=status.HTTP_200_OK)

    return Response({"error": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)


    
@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated, IsSuperAdmin])
def create_user(request):
    serializer = UserSerializer(data=request.data)

    if serializer.is_valid():
        # Create the user but don't save password yet
        password = serializer.validated_data.pop('password')
        user = User.objects.create(
            username=serializer.validated_data['username'],
            email=serializer.validated_data.get('email', ''),
            first_name=serializer.validated_data.get('first_name', ''),
            last_name=serializer.validated_data.get('last_name', ''),
            is_superuser=serializer.validated_data.get('is_superuser', False),
            is_staff=serializer.validated_data.get('is_staff', False)
        )
        
        # Set password separately for proper hashing
        user.set_password(password)
        user.save()
        
        # Generate token
        token, created = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'user': UserSerializer(user).data}, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated, IsSuperAdmin])  # Only superadmins can update users
def update_user(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    data = request.data
    serializer = UserSerializer(user, data=data, partial=True, context={'request': request})

    if serializer.is_valid():
        serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_200_OK)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])  # Any logged-in user can update their own password
def update_own_password(request):
    user = request.user
    old_password = request.data.get("old_password")
    new_password = request.data.get("new_password")

    if not old_password or not new_password:
        return Response({"error": "Old password and new password are required"}, status=status.HTTP_400_BAD_REQUEST)

    if not user.check_password(old_password):
        return Response({"error": "Old password is incorrect"}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)
    user.save()
    return Response({"message": "Password updated successfully"}, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated]) 
def user_list(request):
    users = User.objects.all()
    serializer = UserSerializer(users, many=True, context={'request': request})
    return Response(serializer.data)

@api_view(['DELETE'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated, IsSuperAdmin]) 
def user_delete(request, id):
    try:
        users = User.objects.get(id=id)
    except User.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    users.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['PATCH'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated, IsSuperAdmin])  # Only superadmins can toggle user activation
def toggle_user_activation(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    user.is_active = not user.is_active
    user.save()
    return Response({"message": f"User {'activated' if user.is_active else 'deactivated'} successfully"}, status=status.HTTP_200_OK)

#get items without pagination
@api_view(['GET'])
def get_all_items(request):
    items = Item.objects.all()
    serializer = ItemSerializer(items, many=True,  context={'request': request})
    return Response(serializer.data)

#get items with pagination
@api_view(['GET'])
def item_list_all(request):
    paginator = Pagination()
    items = Item.objects.all()
    result_page = paginator.paginate_queryset(items, request)
    serializer = ItemSerializer(result_page, many=True, context={'request': request})
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
def item_list_detail(request, id):
    try:
        item = Item.objects.get(itemID=id)
    except Item.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    serializer = ItemSerializer(item, context={'request': request})
    return Response(serializer.data)

@api_view(['GET'])
def item_list_search(request):
    search_term = request.query_params.get('search', '')
    
    # Use Q objects for more complex searching
    items = Item.objects.filter(
        Q(itemID__icontains=search_term) | 
        Q(itemName__icontains=search_term) | 
        Q(classificationID__classification__icontains=search_term)
    )
    
    paginator = Pagination()
    result_page = paginator.paginate_queryset(items, request)
    serializer = ItemSerializer(result_page, many=True, context={'request': request})
    return paginator.get_paginated_response(serializer.data)

@api_view(['POST'])
def item_create(request):
    serializer = ItemSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny 

@csrf_exempt
@api_view(['POST'])
def item_bulk_create(request):
    if request.method == 'POST':
        items_data = request.data

        # Added logging for debugging
        print("Received bulk create request")
        print("Request data:", items_data)

        created_items = []
        errors = []
        print("Invalid items:", errors)  # Print any errors that occur during processing

        for item_data in items_data:
            item_id = item_data.get('itemID')
            if Item.objects.filter(itemID=item_id).exists():
                continue  # Skip existing items

            serializer = ItemSerializer(data=item_data)
            if serializer.is_valid():
                serializer.save()
                created_items.append(serializer.data)
            else:
                errors.append(serializer.errors)

        if errors:
            return Response({"created_items": created_items, "errors": errors}, status=status.HTTP_207_MULTI_STATUS)
        return Response(created_items, status=status.HTTP_201_CREATED)

    # Explicitly handle other methods
    return Response({"detail": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['PUT'])
def item_update(request, id):
    try:
        item = Item.objects.get(itemID=id)
    except Item.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    serializer = ItemSerializer(item, data=request.data, partial=True, context={'request': request})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def item_delete(request, id):    
    try:
        items = Item.objects.get(itemID=id)
    except Item.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    items.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)
            
@api_view(['GET'])
def get_beginning_bal(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    measurement_id = request.query_params.get('measurementID')

    beginning_bal = BeginningBalance.objects.all()
    
    if month:
        beginning_bal = beginning_bal.filter(created_at__month=month)
    if year:
        beginning_bal = beginning_bal.filter(created_at__year=year)
    if measurement_id:
        beginning_bal = beginning_bal.filter(measurementID=measurement_id)
        
    paginator = Pagination()
    result_page = paginator.paginate_queryset(beginning_bal, request)
    serializer = BeginningBalanceSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
def search_beginning_bal(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    measurement_id = request.query_params.get('measurementID')
    search_term = request.query_params.get('search', '')
    paginator = Pagination()
    
    beginning_bal = BeginningBalance.objects.filter(
        Q(itemID__icontains=search_term) |
        Q(itemName__icontains=search_term)
    )
    
    # Use Q objects for more complex searching
    if month:
        beginning_bal = beginning_bal.filter(created_at__month=month)
    if year:
        beginning_bal = beginning_bal.filter(created_at__year=year)
    if measurement_id:
        beginning_bal = beginning_bal.filter(measurementID=measurement_id)
        
    result_page = paginator.paginate_queryset(beginning_bal, request)
    serializer = BeginningBalanceSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@csrf_exempt
def copy_items_to_balance(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    current_month = timezone.now().month
    current_year = timezone.now().year

    # Check if BeginningBalance entries for the current month and year already exist
    if BeginningBalance.objects.filter(created_at__month=current_month, created_at__year=current_year).exists():
        return JsonResponse({'error': 'Items have already been copied for this month'}, status=400)

    items = Item.objects.all()

    # Create BeginningBalance entries in bulk
    beginning_balances = [
        BeginningBalance(
            itemID=item.itemID,
            itemName=item.itemName,
            measurementID=item.measurementID,
            itemQuantity=item.itemQuantity,
            unitCost=item.unitCost,
            totalCost=item.itemQuantity * item.unitCost,
            created_at=timezone.now()
        ) for item in items
    ]

    # Bulk insert for efficiency
    BeginningBalance.objects.bulk_create(beginning_balances)

    return JsonResponse({'message': 'Items copied successfully!'}, status=201)

@api_view(['POST'])
def beginning_balance_bulk_create(request):
    if request.method == 'POST':
        balances_data = request.data

        created_balances = []
        errors = []

        for balance_data in balances_data:
            serializer = BeginningBalanceSerializer(data=balance_data)
            if serializer.is_valid():
                serializer.save()
                created_balances.append(serializer.data)
            else:
                errors.append(serializer.errors)

        if errors:
            return Response({"created_balances": created_balances, "errors": errors}, status=status.HTTP_207_MULTI_STATUS)
        return Response(created_balances, status=status.HTTP_201_CREATED)

    return Response({"detail": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET'])
def classification_list_all(request):
    classifications = Classification.objects.all()
    serializer = ClassificationSerializer(classifications, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def measurement_list_all(request):
    measurements = Measurement.objects.all()
    serializer = MeasurementSerializer(measurements, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def section_list_all(request):
    sections = Section.objects.all()
    serializer = SectionSerializer(sections, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def purpose_list_all(request):
    purposes = Purpose.objects.all()
    serializer = PurposeSerializer(purposes, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def area_list_all(request):
    area = Area.objects.all()
    serializer = AreaSerializer(area, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def get_all_transaction_details(request):
    transactionDet = TransactionDetails.objects.all()
    serializer = TransactionDetailsSerializer(transactionDet, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def get_transaction_details(request,id):
    try:
        transactionDet = TransactionDetails.objects.get(transactionDetailsID=id)
    except TransactionDetails.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    serializer = TransactionDetailsSerializer(transactionDet, context={'request': request})
    return Response(serializer.data)

@api_view(['GET'])
def get_all_transaction_product(request):
    transactionProd = TransactionProduct.objects.all()
    serializer = TransactionProductSerializer(transactionProd, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def get_transaction_product(request, id, detailID):
    try:
        transactionProd = TransactionProduct.objects.get(transactionDetailsID_id=detailID, transactionProductID = id)
    except TransactionProduct.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    serializer = TransactionProductSerializer(transactionProd, context={'request': request})
    return Response(serializer.data)

@api_view(['GET'])
def get_transaction_by_itemID(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    item_id = request.query_params.get('itemID')
    
    transaction_history = TransactionProduct.objects.all()
    
    if month:
        transaction_history = TransactionProduct.objects.filter(created_at__month = month)
    if year:
        transaction_history = TransactionProduct.objects.filter(created_at__year = year)
    if item_id:
        transaction_history = TransactionProduct.objects.filter(itemID_id = item_id)

    serializer = TransactionProductSerializer(transaction_history, many=True, context={'request':request})
    return Response(serializer.data)

def calculate_week(date_str):
    today = now().date()
    week_number = (today.day - 1) // 7 + 1 
    return int(week_number)

@api_view(['POST'])
def transaction_detail_create(request):
    if request.method == 'POST':
        serializer = TransactionDetailsSerializer(data=request.data)
        if serializer.is_valid():
            transaction_details = serializer.save()
            transaction_details.week = calculate_week(transaction_details.date)
            transaction_details.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
def transaction_detail_update(request, id):
    try:
        transactionDet = TransactionDetails.objects.get(transactionDetailsID=id)
    except TransactionDetails.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = TransactionDetailsSerializer(transactionDet, data=request.data, partial=True, context={'request': request})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def transaction_detail_delete(request, id):
    try:
        transactionDet = TransactionDetails.objects.get(transactionDetailsID=id)
    except TransactionDetails.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    transactionDet.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['POST'])
def transaction_product_create(request):
    serializer = TransactionProductSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
def transaction_product_update(request, id, detailID):
    try:
        transactionProd = TransactionProduct.objects.get(transactionProductID=id, transactionDetailsID_id = detailID)
        
    except TransactionProduct.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = TransactionProductSerializer(transactionProd, data=request.data, partial=True, context={'request': request})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def transaction_product_delete(request, id, detailID):
    try:
        transactionProd = TransactionProduct.objects.get(transactionProductID=id, transactionDetailsID_id = detailID)
    except TransactionDetails.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    transactionProd.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)
# end of transaction

@api_view(['GET'])
def get_running_balance(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    measurement_id = request.query_params.get('measurementID')
    remarks = request.query_params.get('remarks')
    available_only = request.query_params.get('available_only')
    
    running_balances = RunningBalance.objects.all()

    if month:
        running_balances = running_balances.filter(created_at__month=month)
    if year:
        running_balances = running_balances.filter(created_at__year=year)
    if measurement_id:
        running_balances = running_balances.filter(measurementID=measurement_id)
    if remarks:
        running_balances = running_balances.filter(remarks=remarks)
    if available_only == 'true':
        running_balances = running_balances.filter(itemQuantity__gt=0)
        
    paginator = Pagination()
    result_page = paginator.paginate_queryset(running_balances, request)
    serializer = RunningBalanceSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
def search_running_balance(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    measurement_id = request.query_params.get('measurementID')
    available_only = request.query_params.get('available_only')
    search_term = request.query_params.get('search', '')
    
    paginator = Pagination()
    
    running_balances = RunningBalance.objects.filter(
        Q(itemName__icontains=search_term)
    )

    if month:
        running_balances = running_balances.filter(created_at__month=month)
    if year:
        running_balances = running_balances.filter(created_at__year=year)
    if measurement_id:
        running_balances = running_balances.filter(measurementID=measurement_id)
    if available_only == 'true':
        running_balances = running_balances.filter(itemQuantity__gt=0)
        
    
    result_page = paginator.paginate_queryset(running_balances, request)
    serializer = RunningBalanceSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

import logging
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from django.db.models import Prefetch

logger = logging.getLogger(__name__)

@csrf_exempt 
@api_view(['POST']) 
def create_update_runbal(request): 
    current_month = timezone.now().month 
    current_year = timezone.now().year 
    month, last = (12, current_year) if current_month == 1 else (current_month, current_year)

    # Use select_related to reduce database queries for related models
    items = Item.objects.all().select_related('measurementID').order_by('itemID')
    
    # Pre-fetch all beginning balances for the previous month in one query
    beginning_balances = {
        str(bb.itemID): bb.itemQuantity 
        for bb in BeginningBalance.objects.filter(
            created_at__year=last,
            created_at__month=month,
        )
    }

    # Prefetch all existing running balances for the current month
    existing_balances = {}
    for rb in RunningBalance.objects.filter(
        created_at__month=current_month,
        created_at__year=current_year
    ):
        # Convert to string to ensure consistent lookup
        existing_balances[str(rb.itemID_id)] = rb
    
    # Fetch aggregated transaction data for all items at once using Django's aggregation
    transaction_sums = {}
    for item_data in TransactionProduct.objects.filter(
        created_at__month=current_month,
        created_at__year=current_year
    ).values('itemID').annotate(
        purchasedFromSupp_sum=Sum('purchasedFromSupp'),
        returnToSupplier_sum=Sum('returnToSupplier'),
        transferFromWH_sum=Sum('transferFromWH'),
        transferToWH_sum=Sum('transferToWH'),
        issuedQty_sum=Sum('issuedQty'),
        returnedQty_sum=Sum('returnedQty'),
    ):
        itemID = item_data['itemID']
        transaction_sums[itemID] = {
            'purchasedFromSupp_sum': item_data['purchasedFromSupp_sum'] or 0,
            'returnToSupplier_sum': item_data['returnToSupplier_sum'] or 0,
            'transferFromWH_sum': item_data['transferFromWH_sum'] or 0,
            'transferToWH_sum': item_data['transferToWH_sum'] or 0,
            'issuedQty_sum': item_data['issuedQty_sum'] or 0,
            'returnedQty_sum': item_data['returnedQty_sum'] or 0,
        }
    
    # Pre-fetch consumption data grouped by item and week
    consumption_data = {}
    for consumption in MonthlyConsumption.objects.filter(
        date__month=current_month,
        date__year=current_year
    ).values('itemID', 'week').annotate(weekly_count=Sum('week')):
        itemID = consumption['itemID']
        
        if itemID not in consumption_data:
            consumption_data[itemID] = []
            
        consumption_data[itemID].append({
            'week': consumption['week'],
            'count': consumption['weekly_count'] or 0
        })
    
    # Create batch operations lists
    balances_to_create = []
    balances_to_update = []
    
    for item in items:
        itemID = item.itemID  # or item.pk depending on your model
        
        # Get or initialize running balance
        if itemID in existing_balances:
            running_balance = existing_balances[itemID]
        else:
            # print(f"ItemID: {itemID}, BeginningBalance: {beginning_balances.get(str(itemID), 'NOT FOUND')}")
            
            running_balance = RunningBalance(
                itemID=item,
                itemName=item.itemName, 
                beginningBalance=beginning_balances.get(str(item.itemID), 0),
                measurementID=item.measurementID, 
                itemQuantity=item.itemQuantity, 
                unitCost=item.unitCost, 
                purchasedFromSupp=0, 
                returnToSupplier=0, 
                transferFromWH=0, 
                transferToWH=0, 
                issuedQty=0, 
                returnedQty=0, 
                consumption=0, 
                totalCost=0,
                created_at=timezone.now()
            )

        # Set beginning balance
        running_balance.beginningBalance = beginning_balances.get(str(itemID), 0)
        
        # Update with transaction sums if available
        sums = transaction_sums.get(itemID, {
            'purchasedFromSupp_sum': 0,
            'returnToSupplier_sum': 0,
            'transferFromWH_sum': 0,
            'transferToWH_sum': 0,
            'issuedQty_sum': 0,
            'returnedQty_sum': 0,
        })
        
        # Update summary fields
        running_balance.itemQuantity = item.itemQuantity
        running_balance.beginningBalance = beginning_balances.get(str(running_balance.itemID_id), 0)
        running_balance.purchasedFromSupp = sums['purchasedFromSupp_sum']
        running_balance.returnToSupplier = sums['returnToSupplier_sum']
        running_balance.transferFromWH = sums['transferFromWH_sum']
        running_balance.transferToWH = sums['transferToWH_sum']
        running_balance.issuedQty = sums['issuedQty_sum']
        running_balance.returnedQty = sums['returnedQty_sum']
        running_balance.consumption = running_balance.returnedQty - running_balance.issuedQty
        running_balance.unitCost = item.unitCost
        running_balance.totalCost = item.unitCost * running_balance.itemQuantity
        
        # Set remarks based on consumption data
        item_consumption = consumption_data.get(itemID, [])
        
        if not item_consumption:
            running_balance.remarks = "Non-Moving"
        else:
            # Calculate total consumption and active weeks
            total_consumption = sum(week_data['count'] for week_data in item_consumption)
            active_weeks = len([week_data for week_data in item_consumption if week_data['count'] > 0])
            
            weekly_average = total_consumption / active_weeks if active_weeks > 0 else 0
            
            if weekly_average == 0:
                running_balance.remarks = "Non-Moving"
            elif weekly_average < 10:
                running_balance.remarks = "Slow Moving"
            else:
                running_balance.remarks = "Fast Moving"
                
        # print(f"ItemID: {itemID}, BeginningBalance: {beginning_balances.get(str(itemID), 'NOT FOUND UPDATE')}")
        
        # Add to appropriate batch list
        if itemID in existing_balances:
            balances_to_update.append(running_balance)
            running_balance.save()
        else:
            balances_to_create.append(running_balance)
    
    # Batch create and update operations
    if balances_to_create:
        RunningBalance.objects.bulk_create(balances_to_create)
    
    if balances_to_update:
        RunningBalance.objects.bulk_update(
            balances_to_update,
            fields=['beginningBalance', 'itemQuantity', 'purchasedFromSupp', 
                    'returnToSupplier', 'transferFromWH', 'transferToWH', 
                    'issuedQty', 'returnedQty', 'consumption', 'unitCost', 
                    'totalCost', 'remarks']
        )

    return Response({"detail": "Running balance processed successfully for all items."})

@api_view(['GET'])
def get_monthly_consumption(request):
    # Get query parameters
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    section_id = request.query_params.get('sectionID')
    
    # Filter MonthlyConsumption based on the provided parameters
    monthly_consumption = MonthlyConsumption.objects.all()

    if month:
        monthly_consumption = monthly_consumption.filter(date__month=month)
    if year:
        monthly_consumption = monthly_consumption.filter(date__year=year)
    if section_id:
        monthly_consumption = monthly_consumption.filter(sectionID=section_id)

    paginator = Pagination()
    
    # Order by week to ensure proper weekly breakdown
    monthly_consumption = monthly_consumption.order_by('week')

    result_page = paginator.paginate_queryset(monthly_consumption, request)
    serializer = MonthlyConsumptionSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
def search_monthly_consumption(request):
    # Get query parameters
    month = request.query_params.get('month')
    year = request.query_params.get('year')
    section_id = request.query_params.get('sectionID')
    search_term = request.query_params.get('search', '')
    
    # Filter MonthlyConsumption based on the provided parameters
    monthly_consumption = MonthlyConsumption.objects.filter(
        Q(itemName__icontains=search_term) | 
        Q(date__icontains=search_term) 
    )

    if month:
        monthly_consumption = monthly_consumption.filter(date__month=month)
    if year:
        monthly_consumption = monthly_consumption.filter(date__year=year)
    if section_id:
        monthly_consumption = monthly_consumption.filter(sectionID=section_id)

    paginator = Pagination()
    
    # Order by week to ensure proper weekly breakdown
    monthly_consumption = monthly_consumption.order_by('week')

    result_page = paginator.paginate_queryset(monthly_consumption, request)
    # Serialize the filtered data
    serializer = MonthlyConsumptionSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
def export_consumption_to_excel(request):
    # Extract filter parameters
    section_id = request.GET.get('sectionID') or request.query_params.get('sectionID')
    month = request.GET.get('month') or request.query_params.get('month')
    year = request.GET.get('year') or request.query_params.get('year')
    
    # Convert month number to month name
    month_names = {
        '1': 'January', '2': 'February', '3': 'March', 
        '4': 'April', '5': 'May', '6': 'June', 
        '7': 'July', '8': 'August', '9': 'September', 
        '10': 'October', '11': 'November', '12': 'December'
    }
    month_name = month_names.get(str(month), 'All Months')
    
    # Query the consumption data based on the parameters
    consumption_data = MonthlyConsumption.objects.all()
    
    # Apply filters if provided
    if section_id and int(section_id) != 0:  # Export all if sectionID is 0
        consumption_data = consumption_data.filter(sectionID=section_id)
    if month:
        consumption_data = consumption_data.filter(date__month=month)
    if year:
        consumption_data = consumption_data.filter(date__year=year)
    
    # Convert the queryset to a DataFrame
    df = pd.DataFrame(list(consumption_data.values(
        'date', 
        'week', 
        'sectionName',
        'itemID', 
        'itemName', 
        'consumption', 
        'cost', 
        'total'
    )))
    
    # Create filename based on filters
    filename_parts = []
    section = None
    if section_id and int(section_id) != 0:
        section = Section.objects.filter(sectionID=section_id).first()
        if section:
            filename_parts.append(f"section_{section.sectionName}")
    else:
        filename_parts.append("all_sections")
    if month:
        filename_parts.append(f"Month_{month_name}")
    if year:
        filename_parts.append(f"year_{year}")
    
    filename = "_".join(filename_parts) if filename_parts else "all_consumption"
    filename += "_report.xlsx"
    
    # Path to the template and logo
    template_path = os.path.join(settings.BASE_DIR, 'Document Format', 'MONTHLY CONSUMPTION SECTION.xlsx')
    logo_path = os.path.join(settings.BASE_DIR, 'Document Format', 'CABWAD LOGO.png')
    
    # Create a temporary file to save the Excel file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        temp_file_path = tmp_file.name
        
        # Load the template workbook
        wb = load_workbook(template_path)
        
        # Get the first worksheet
        ws = wb.active
        
        # Insert logo
        if os.path.exists(logo_path):
            img = Image(logo_path)
            # Adjust size and position as needed
            img.width = 120
            img.height = 120
            img.anchor = 'A1'
            ws.add_image(img)
        
        # Improved placeholder replacement
        for row in range(1, 10):  # Check first 10 rows
            for col in range(1, 10):  # Check first 10 columns
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    # Replace multiple placeholders in the same cell
                    new_value = cell.value.replace('(section)', section.sectionName if section else 'All Sections')
                    new_value = new_value.replace('(month)', month_name)
                    new_value = new_value.replace('(year)', year or 'All Years')
                    cell.value = new_value
        
        # Write DataFrame headers and data starting from row 9
        headers = ['Date', 'Week', 'Section Name', 'Item ID', 'Item Name', 'Consumption', 'Cost', 'Total']
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=c_idx, value=header)
            
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=9, column=c_idx, value=header)
        
        # Write data rows
        for r_idx, row in enumerate(df.values.tolist(), start=10):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Save the workbook
        wb.save(temp_file_path)
    
    # Serve the file for download
    response = FileResponse(open(temp_file_path, 'rb'), as_attachment=True, filename=filename)
    
    # Clean up the temporary file after serving
    response['Cleanup-Temp-File'] = temp_file_path  # Custom header for cleanup
    return response

# @csrf_exempt  # Remove in production; use proper authentication
# def xlsm_to_json(request):
#     if request.method == "POST" and request.FILES.get("file"):
#         excel_file = request.FILES["file"]
#         sheet_name = request.POST.get("sheet_name", "Beginning Balance")  # Default to "Sheet1"
        
#         # Save file temporarily, overwrite if exists
#         file_path = f"temp/{excel_file.name}"
#         if default_storage.exists(file_path):
#             default_storage.delete(file_path)
#         file_path = default_storage.save(file_path, excel_file)

#         try:
#             # Read the specified sheet from the .xlsm file
#             df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            
#             column_mapping = {
#                 "Item ID": "itemID",
#                 "PRODUCT NAME": "itemName",
#                 "UNIT OF MEASURE": "measurementName",
#                 "Available Stocks": "itemQuantity",
#                 "Ave. Unit Cost": "unitCost"
#             } 
            
#             # Specify the columns you want to extract
#             selected_columns = ["Item ID","PRODUCT NAME", "UNIT OF MEASURE", "Available Stocks", "Ave. Unit Cost"]
#             df_selected = df[selected_columns].rename(columns=column_mapping)

#             # Trim whitespace from string columns
#             df_selected = df_selected.applymap(lambda x: x.strip() if isinstance(x, str) else x)

#             # Replace NaN with None (which becomes null in JSON)
#             df_selected = df_selected.where(pd.notnull(df_selected), 0)

#             # Convert to JSON and trim all string values
#             json_data = df_selected.to_dict(orient="records")
#             for item in json_data:
#                 for key, value in item.items():
#                     if isinstance(value, str):
#                         item[key] = value.strip()

#             response = JsonResponse(json_data, safe=False)
#             default_storage.delete(file_path)
#             return response

#         except ValueError as ve:
#             return JsonResponse({"error": f"Sheet '{sheet_name}' not found"}, status=400)
#         except KeyError as ke:
#             return JsonResponse({"error": f"Missing columns: {ke}"}, status=400)
#         except Exception as e:
#             return JsonResponse({"error": str(e)}, status=500)

#     return JsonResponse({"error": "Invalid request"}, status=400)

@csrf_exempt  # Remove in production; use proper authentication
def xlsm_to_json(request):
    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        sheet_name = request.POST.get("sheet_name", "Masterlist")  # Default to "Sheet1"
        additional_sheet_name = request.POST.get("additional_sheet", "Beginning Balance")
        
        # Save file temporarily, overwrite if exists
        file_path = f"temp/{excel_file.name}"
        if default_storage.exists(file_path):
            default_storage.delete(file_path)
        file_path = default_storage.save(file_path, excel_file)

        try:
            # Read the specified sheet from the .xlsm file
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            df_additional = pd.read_excel(file_path, sheet_name=additional_sheet_name, engine="openpyxl") 
            
            column_mapping = {
                "PRODUCT NAME": "itemName",
                "CLASSIFICATION": "classificationName"
            } 
            
            # Specify the columns you want to extract
            selected_columns = ["PRODUCT NAME", "CLASSIFICATION"]
            df_selected = df[selected_columns].rename(columns=column_mapping)

            additional_col_map = {
                "PRODUCT NAME": "itemName", 
                "UNIT OF MEASURE": "measurementName",
                "Available Stocks": "itemQuantity",
                "Ave. Unit Cost": "unitCost"
            }
            # Extract Unit of Measurement from the additional sheet
            additional_columns = ["PRODUCT NAME", "UNIT OF MEASURE", "Available Stocks", "Ave. Unit Cost"]
            df_additional_selected = df_additional[additional_columns].rename(columns=additional_col_map)

            # Merge the dataframes on "itemName"
            df_merged = pd.merge(df_selected, df_additional_selected, on="itemName", how="left")

            # Trim whitespace from string columns
            df_merged = df_merged.applymap(lambda x: x.strip() if isinstance(x, str) else x)

            # Ensure itemQuantity is 0.00 if it's 0
            df_merged['itemQuantity'] = df_merged['itemQuantity'].apply(lambda x: 0.00 if x == 0 else x)
            df_merged['unitCost'] = df_merged['unitCost'].apply(lambda x: 0.00 if x == 0 else x)
            # Replace NaN with None (which becomes null in JSON)
            df_merged = df_merged.where(pd.notnull(df_merged), "N/A")



            # Convert to JSON and trim all string values
            json_data = df_merged.to_dict(orient="records")
            for item in json_data:
                for key, value in item.items():
                    if isinstance(value, str):
                        item[key] = value.strip()

            response = JsonResponse(json_data, safe=False)
            default_storage.delete(file_path)
            return response

        except ValueError as ve:
            return JsonResponse({"error": f"Sheet '{sheet_name}' not found"}, status=400)
        except KeyError as ke:
            return JsonResponse({"error": f"Missing columns: {ke}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)

from django.http import FileResponse
from rest_framework.response import Response
from rest_framework.decorators import api_view
from django.views.decorators.csrf import csrf_exempt
import os
import logging
from ims.utils import generate_reports_doc
import tempfile

logger = logging.getLogger(__name__)

@csrf_exempt
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_report_doc(request, year, month):
    try:
        # Validate input parameters
        try:
            year = int(year)
            month = int(month)
            
            # Ensure valid month range
            if month < 1 or month > 12:
                return Response({"error": "Invalid month. Must be between 1 and 12."}, status=400)
        except ValueError:
            return Response({"error": "Invalid year or month format"}, status=400)

        # Query the report by year and month
        try:
            report = MonthlyConsumptionTotal.objects.get(
                updated_at__year=year, 
                updated_at__month=month
            )
        except MonthlyConsumptionTotal.DoesNotExist:
            logger.warning(f"Report not found for year {year}, month {month}")
            return Response({
                "error": "Report not found for the specified year and month",
                "details": {
                    "year": year,
                    "month": month
                }
            }, status=404)
        
        # Generate the report
        try:
            doc_path = generate_reports_doc(report)
        except Exception as doc_gen_error:
            logger.error(f"Failed to generate report: {str(doc_gen_error)}", exc_info=True)
            return Response({
                "error": "Failed to generate report document",
                "details": str(doc_gen_error)
            }, status=500)
        
        # Determine file to serve (DOCX)
        if os.path.exists(doc_path):
            file_path = doc_path
            filename = f"report_{year}_{month}.docx"
            content_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        else:
            logger.error("No valid file generated")
            return Response({
                "error": "No valid report file found",
                "details": {
                    "docx_path": doc_path
                }
            }, status=500)
        
        # Serve the file for download
        try:
            response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename, content_type=content_type)
            
            # Basic caching headers
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            
            logger.info(f"Report downloaded: Year {year}, Month {month}, Format: DOCX")
            
            return response
        
        except Exception as file_error:
            logger.error(f"Error serving file: {str(file_error)}", exc_info=True)
            return Response({
                "error": "Unexpected error serving file",
                "details": str(file_error)
            }, status=500)
    
    except Exception as unexpected_error:
        logger.critical(f"Unexpected error in report download: {str(unexpected_error)}", exc_info=True)
        return Response({
            "error": "An unexpected error occurred",
            "details": str(unexpected_error)
        }, status=500)
